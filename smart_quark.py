import os
import re
import sys
import io
import time
import random
import json
import logging
import threading
import requests
import openpyxl
import tkinter as tk
from tkinter import ttk
from tkinter import messagebox
from tkinter import filedialog
from datetime import datetime

# 强制标准输出采用 UTF-8 编码，若为 None (如 PyInstaller --noconsole 模式) 则重定向到 devnull
if sys.stdout is None:
    sys.stdout = open(os.devnull, 'w')
elif hasattr(sys.stdout, 'buffer'):
    sys.stdout = io.TextIOWrapper(sys.stdout.buffer, encoding='utf-8')

if sys.stderr is None:
    sys.stderr = open(os.devnull, 'w')
elif hasattr(sys.stderr, 'buffer'):
    sys.stderr = io.TextIOWrapper(sys.stderr.buffer, encoding='utf-8')


DOMAIN_TLDS = "com|cn|net|org|cc|info|xyz|me|top|vip|co|club|wang|run|pub|love|icu|work|site|online|link|fun|mobi|biz|win|space"
DOMAIN_TLD_SET = set(DOMAIN_TLDS.split("|"))
DOMAIN_PATTERN = r'(?:[a-zA-Z0-9](?:[a-zA-Z0-9-]{0,61}[a-zA-Z0-9])?\.)+(?:' + DOMAIN_TLDS + r')'
SEPARATORS = r' \-_+_\|｜'
RESOURCE_EXTS = {
    '.mp4', '.mkv', '.avi', '.mov', '.wmv', '.flv', '.f4v', '.rmvb',
    '.mp3', '.wav', '.wma', '.aac', '.flac', '.m4a',
    '.zip', '.rar', '.7z', '.tar', '.gz',
    '.pdf', '.epub', '.mobi', '.azw3', '.docx', '.xlsx', '.pptx',
    '.exe', '.dmg', '.pkg', '.apk', '.iso'
}
AD_EXTS = {'.jpg', '.jpeg', '.png', '.gif', '.bmp', '.html', '.url', '.lnk'}
PROMO_CHARS = ['群', '领', '众', '信', '微', '扫', '码', '加', '粉', '福利']


def clean_filename(filename, keywords):
    base_name, ext = os.path.splitext(filename)
    if ext[1:].lower() in DOMAIN_TLD_SET and re.search(DOMAIN_PATTERN, filename, re.IGNORECASE):
        base_name, ext = filename, ""
    
    # 自动提取并清理文件名中的网址/域名及其所在括号
    domain_brackets = [
        r'【[^】]*?' + DOMAIN_PATTERN + r'[^】]*?】',
        r'\[[^\]]*?' + DOMAIN_PATTERN + r'[^\]]*?\]',
        r'\([^)]*?' + DOMAIN_PATTERN + r'[^)]*?\)',
        r'（[^）]*?' + DOMAIN_PATTERN + r'[^）]*?）',
        r'\{[^}]*?' + DOMAIN_PATTERN + r'[^}]*?\}'
    ]
    for pattern in domain_brackets:
        base_name = re.sub(pattern, '', base_name, flags=re.IGNORECASE)
    sep_class = '[' + SEPARATORS + r']+'
    base_name = re.sub(sep_class + DOMAIN_PATTERN + r'(?=$|' + sep_class + r')', '', base_name, flags=re.IGNORECASE)
    base_name = re.sub(r'^' + DOMAIN_PATTERN + sep_class, '', base_name, flags=re.IGNORECASE)
    if re.fullmatch(DOMAIN_PATTERN, base_name, flags=re.IGNORECASE):
        base_name = ''

    for kw in keywords:
        if not kw:
            continue
        escaped_kw = re.escape(kw)
        
        # 1. 匹配括号包裹的关键字：【...kw...】、[...kw...] 等
        brackets = [
            r'【[^】]*?' + escaped_kw + r'[^】]*?】',
            r'\[[^\]]*?' + escaped_kw + r'[^\]]*?\]',
            r'\([^)]*?' + escaped_kw + r'[^)]*?\)',
            r'（[^）]*?' + escaped_kw + r'[^）]*?）',
            r'\{[^}]*?' + escaped_kw + r'[^}]*?\}'
        ]
        for pattern in brackets:
            base_name, count = re.subn(pattern, '', base_name, flags=re.IGNORECASE)
            if count > 0:
                break
        
        # 2. 匹配分隔符开头的关键字及其后续字符，如： -公众号-xxx, _微信xxx 等
        pattern_sep = r'[ \-_+_\|｜]+[^ \-_+_\|｜]*?' + escaped_kw + r'[^ \-_+_\|｜]*'
        base_name, count = re.subn(pattern_sep, '', base_name, flags=re.IGNORECASE)
        
        # 3. 直接匹配关键字、冒号及随后的空格，如： 公众号：, 微信:
        if count == 0:
            pattern_colon = escaped_kw + r'[:：\s]*'
            base_name = re.sub(pattern_colon, '', base_name, flags=re.IGNORECASE)
            
    # 清理多余空格和首尾的分隔符
    base_name = re.sub(r'\s+', ' ', base_name)
    base_name = re.sub(r'^[ \-_+_\|｜]+|[ \-_+_\|｜]+$', '', base_name)
    base_name = base_name.strip()
    
    if not base_name:
        return ""
    return base_name + ext


def has_domain_pattern(name):
    return re.search(DOMAIN_PATTERN, name, re.IGNORECASE) is not None


def classify_file_action(name, size, is_dir, blacklist_keywords):
    """Return (action, new_name): action is keep, rename, or delete."""
    is_sensitive = any(kw and kw in name for kw in blacklist_keywords)
    if not is_sensitive and has_domain_pattern(name):
        is_sensitive = True

    ext = os.path.splitext(name)[1].lower()
    if not is_sensitive and not is_dir and ext in AD_EXTS:
        if ext in {'.html', '.url', '.lnk'}:
            is_sensitive = True
        elif size > 0 and size < 150 * 1024 and any(c in name for c in PROMO_CHARS):
            is_sensitive = True
        elif size > 0 and size < 100 * 1024:
            base_name = os.path.splitext(name)[0]
            if re.match(r'^[0-9_\-]+$', base_name):
                is_sensitive = True

    if not is_sensitive:
        return "keep", None

    new_name = clean_filename(name, blacklist_keywords)
    if not is_dir and size < 100 * 1024:
        return "delete", None
    if is_dir or ext in RESOURCE_EXTS:
        return "rename", new_name
    if ext in AD_EXTS:
        return "delete", None

    base = os.path.splitext(name)[0]
    if not new_name or len(base) < 12 or size < 50 * 1024:
        return "delete", None
    return "rename", new_name


def safe_int(value, default=0):
    try:
        return int(value)
    except Exception:
        return default


def sleep_with_stop(seconds, stop_flag=None, interval=0.2, sleep_func=time.sleep):
    remaining = float(seconds)
    while remaining > 0:
        if stop_flag and stop_flag():
            return False
        step = min(interval, remaining)
        sleep_func(step)
        remaining -= step
    return not (stop_flag and stop_flag())


def clean_ads_recursively(transfer, current_fid, book_name, blacklist_keywords, log_func, stop_flag=None):
    if stop_flag and stop_flag():
        return

    contents = transfer.ls_dir(current_fid)
    if not contents or not isinstance(contents, dict) or contents.get("code") != 0:
        return
    
    file_list = contents.get("data", {}).get("list", [])
    for file_item in file_list:
        if stop_flag and stop_flag():
            return

        name = file_item.get("file_name", "")
        fid = file_item.get("fid")
        if not name or not fid:
            continue
        
        # 判断是否是目录
        is_dir = (
            file_item.get("dir") is True or 
            str(file_item.get("dir")) == "1" or 
            str(file_item.get("dir")).lower() == "true" or
            file_item.get("file_type") == "dir"
        )
        
        if is_dir:
            clean_ads_recursively(transfer, fid, book_name, blacklist_keywords, log_func, stop_flag=stop_flag)
            if stop_flag and stop_flag():
                return

        size = safe_int(file_item.get("size") or 0)
        action, new_name = classify_file_action(name, size, is_dir, blacklist_keywords)
        if action == "keep":
            continue
        if action == "rename":
            ext = os.path.splitext(name)[1].lower()
            fallback_name = f"已清洗文件夹_{fid}" if is_dir else f"已清洗资源_{fid}{ext}"
            target_name = new_name if new_name else fallback_name
            if target_name == name:
                continue
            rename_res = transfer.rename(fid, target_name)
            if rename_res and rename_res.get("code") == 0:
                label = "文件夹重命名" if is_dir else "重命名"
                log_func(f"资源 [{book_name}] {label}：[{name}] -> [{target_name}]")
            else:
                log_func(f"资源 [{book_name}] 重命名失败 [{name}]: {rename_res}")
        elif action == "delete":
            del_res = transfer.delete([fid])
            if del_res and del_res.get("code") == 0:
                log_func(f"资源 [{book_name}] 广告文件已删除：[{name}]")
            else:
                log_func(f"资源 [{book_name}] 广告文件删除失败 [{name}]: {del_res}")


def clean_baidu_ads_recursively(transfer, current_path, book_name, blacklist_keywords, log_func, stop_flag=None):
    if stop_flag and stop_flag():
        return

    file_list = transfer.list_directory(current_path)
    if not file_list:
        return
    
    for file_item in file_list:
        if stop_flag and stop_flag():
            return

        name = file_item.get("server_filename", "")
        path = file_item.get("path")
        if not name or not path:
            continue
        
        is_dir = (
            file_item.get("isdir") == 1 or
            str(file_item.get("isdir")) == "1"
        )
        
        if is_dir:
            clean_baidu_ads_recursively(transfer, path, book_name, blacklist_keywords, log_func, stop_flag=stop_flag)
            if stop_flag and stop_flag():
                return

        size = safe_int(file_item.get("size") or 0)
        action, new_name = classify_file_action(name, size, is_dir, blacklist_keywords)
        if action == "keep":
            continue
        if action == "rename":
            ext = os.path.splitext(name)[1].lower()
            fs_id = file_item.get("fs_id")
            fallback_name = f"已清洗文件夹_{fs_id}" if is_dir else f"已清洗资源_{fs_id}{ext}"
            target_name = new_name if new_name else fallback_name
            if target_name == name:
                continue
            rename_res = transfer.rename(path, target_name)
            if rename_res and rename_res.get("errno") == 0:
                label = "文件夹重命名" if is_dir else "重命名"
                log_func(f"百度资源 [{book_name}] {label}：[{name}] -> [{target_name}]")
            else:
                log_func(f"百度资源 [{book_name}] 重命名失败 [{name}]: {rename_res}")
        elif action == "delete":
            del_res = transfer.delete([path])
            if del_res and del_res.get("errno") == 0:
                log_func(f"百度资源 [{book_name}] 广告文件已删除：[{name}]")
            else:
                log_func(f"百度资源 [{book_name}] 广告文件删除失败 [{name}]: {del_res}")


class QuarkBatchTransfer:
    BASE_URL = "https://drive-pc.quark.cn"
    USER_AGENT = "Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/120.0.0.0 Safari/537.36"

    def __init__(self, cookie: str):
        self.cookie = cookie.strip()
        self.headers = {
            "cookie": self.cookie,
            "content-type": "application/json",
            "user-agent": self.USER_AGENT,
            "origin": "https://pan.quark.cn",
            "referer": "https://pan.quark.cn/"
        }

    def _send_request(self, method, url, **kwargs):
        if "headers" not in kwargs:
            kwargs["headers"] = self.headers
        if "timeout" not in kwargs:
            kwargs["timeout"] = 30
        try:
            response = requests.request(method, url, **kwargs)
            return response
        except Exception as e:
            logging.error(f"网络请求异常: {e}")
            return None

    def get_account_info(self):
        url = "https://pan.quark.cn/account/info"
        querystring = {"fr": "pc", "platform": "pc"}
        resp = self._send_request("GET", url, params=querystring)
        if resp and resp.status_code == 200:
            res_json = resp.json()
            if res_json.get("data"):
                return res_json["data"]
        return None

    def get_fids(self, file_paths):
        url = f"{self.BASE_URL}/1/clouddrive/file/info/path_list"
        querystring = {"pr": "ucpro", "fr": "pc"}
        payload = {"file_path": file_paths, "namespace": "0"}
        resp = self._send_request("POST", url, json=payload, params=querystring)
        if resp and resp.status_code == 200:
            res_json = resp.json()
            if res_json.get("code") == 0:
                return res_json.get("data", [])
        return []

    def mkdir(self, dir_path):
        url = f"{self.BASE_URL}/1/clouddrive/file"
        querystring = {"pr": "ucpro", "fr": "pc", "uc_param_str": ""}
        payload = {
            "pdir_fid": "0",
            "file_name": "",
            "dir_path": dir_path,
            "dir_init_lock": False,
        }
        resp = self._send_request("POST", url, json=payload, params=querystring)
        if resp and resp.status_code == 200:
            res_json = resp.json()
            if res_json.get("code") == 0:
                return res_json.get("data")
        return None

    def get_stoken(self, pwd_id):
        url = f"{self.BASE_URL}/1/clouddrive/share/sharepage/token"
        querystring = {"pr": "ucpro", "fr": "pc"}
        payload = {"pwd_id": pwd_id, "passcode": ""}
        resp = self._send_request("POST", url, json=payload, params=querystring)
        if resp and resp.status_code == 200:
            res_json = resp.json()
            if res_json.get("code") == 0 and res_json.get("data"):
                return res_json["data"].get("stoken")
        return None

    def get_detail(self, pwd_id, stoken):
        list_merge = []
        page = 1
        while True:
            url = f"{self.BASE_URL}/1/clouddrive/share/sharepage/detail"
            querystring = {
                "pr": "ucpro",
                "fr": "pc",
                "pwd_id": pwd_id,
                "stoken": stoken,
                "pdir_fid": "0",
                "force": "0",
                "_page": page,
                "_size": "50",
                "_fetch_banner": "0",
                "_fetch_share": "0",
                "_fetch_total": "1",
                "_sort": "file_type:asc,updated_at:desc",
                "ver": "2",
            }
            resp = self._send_request("GET", url, params=querystring)
            if not resp or resp.status_code != 200:
                break
            res_json = resp.json()
            if res_json.get("code") != 0:
                break
            if res_json.get("data", {}).get("list"):
                list_merge += res_json["data"]["list"]
                page += 1
            else:
                break
            if len(list_merge) >= res_json.get("metadata", {}).get("_total", 0):
                break
        return list_merge

    def save_file(self, fid_list, fid_token_list, to_pdir_fid, pwd_id, stoken):
        url = "https://drive.quark.cn/1/clouddrive/share/sharepage/save"
        querystring = {
            "pr": "ucpro",
            "fr": "pc",
            "uc_param_str": "",
            "app": "clouddrive",
            "__dt": int(random.uniform(1, 5) * 60 * 1000),
            "__t": int(time.time() * 1000),
        }
        payload = {
            "fid_list": fid_list,
            "fid_token_list": fid_token_list,
            "to_pdir_fid": to_pdir_fid,
            "pwd_id": pwd_id,
            "stoken": stoken,
            "pdir_fid": "0",
            "scene": "link",
        }
        resp = self._send_request("POST", url, json=payload, params=querystring)
        if resp and resp.status_code == 200:
            res_json = resp.json()
            if res_json.get("code") == 0 and res_json.get("data"):
                return res_json["data"].get("task_id")
        return None

    def ls_dir(self, pdir_fid):
        list_merge = []
        page = 1
        while True:
            url = f"{self.BASE_URL}/1/clouddrive/file/sort"
            querystring = {
                "pr": "ucpro",
                "fr": "pc",
                "uc_param_str": "",
                "pdir_fid": pdir_fid,
                "_page": page,
                "_size": "50",
                "_fetch_total": "1",
                "_fetch_sub_dirs": "0",
                "_sort": "file_type:asc,updated_at:desc",
                "fetch_all_file": 1,
                "fetch_risk_file_name": 1,
            }
            resp = self._send_request("GET", url, params=querystring)
            if not resp or resp.status_code != 200:
                break
            res_json = resp.json()
            if res_json.get("code") != 0:
                return res_json
            if res_json.get("data", {}).get("list"):
                list_merge += res_json["data"]["list"]
                page += 1
            else:
                break
            if len(list_merge) >= res_json.get("metadata", {}).get("_total", 0):
                break
        return {"code": 0, "data": {"list": list_merge}}

    def delete(self, filelist):
        url = f"{self.BASE_URL}/1/clouddrive/file/delete"
        querystring = {"pr": "ucpro", "fr": "pc", "uc_param_str": ""}
        payload = {"action_type": 2, "filelist": filelist, "exclude_fids": []}
        resp = self._send_request("POST", url, json=payload, params=querystring)
        if resp and resp.status_code == 200:
            return resp.json()
        return None

    def rename(self, fid, new_name):
        url = f"{self.BASE_URL}/1/clouddrive/file/rename"
        querystring = {"pr": "ucpro", "fr": "pc", "uc_param_str": ""}
        payload = {"fid": fid, "file_name": new_name}
        resp = self._send_request("POST", url, json=payload, params=querystring)
        if resp and resp.status_code == 200:
            return resp.json()
        return None

    def share_files(self, fid_list, title):
        url = f"{self.BASE_URL}/1/clouddrive/share"
        querystring = {"pr": "ucpro", "fr": "pc", "uc_param_str": ""}
        payload = {
            "fid_list": fid_list,
            "title": title,
            "url_type": 1,      # 无密公开分享
            "expired_type": 1   # 永久有效
        }
        resp = self._send_request("POST", url, json=payload, params=querystring)
        if resp and resp.status_code == 200:
            res_json = resp.json()
            if res_json.get("code") == 0 and res_json.get("data"):
                return res_json["data"].get("task_id")
        return None

    def query_task(self, task_id, stop_flag=None):
        retry_index = 0
        max_retries = 120  # 最多轮询 120 次（约 2 分钟）
        while retry_index < max_retries:
            if stop_flag and stop_flag():
                logging.info("query_task: 收到停止信号，退出轮询")
                return None
            url = f"{self.BASE_URL}/1/clouddrive/task"
            querystring = {
                "pr": "ucpro",
                "fr": "pc",
                "uc_param_str": "",
                "task_id": task_id,
                "retry_index": retry_index,
                "__dt": int(random.uniform(1, 5) * 60 * 1000),
                "__t": int(time.time() * 1000),
            }
            resp = self._send_request("GET", url, params=querystring)
            if not resp or resp.status_code != 200:
                retry_index += 1
                if not sleep_with_stop(1, stop_flag=stop_flag):
                    return None
                continue
            res_json = resp.json()
            if res_json.get("status") != 200:
                return None
            
            task_status = res_json.get("data", {}).get("status")
            if task_status == 2:  # 成功
                return res_json
            elif task_status == 3:  # 失败
                return None
            else:
                retry_index += 1
                if not sleep_with_stop(1, stop_flag=stop_flag):
                    return None
        logging.warning(f"query_task: 超过最大轮询次数 {max_retries}，task_id={task_id}")
        return None

    def get_share_link(self, share_id):
        url = f"{self.BASE_URL}/1/clouddrive/share/password"
        querystring = {"pr": "ucpro", "fr": "pc", "uc_param_str": ""}
        payload = {"share_id": share_id}
        resp = self._send_request("POST", url, json=payload, params=querystring)
        if resp and resp.status_code == 200:
            res_json = resp.json()
            if res_json.get("code") == 0 and res_json.get("data"):
                return res_json["data"].get("share_url")
        return None


def get_id_from_url(url: str) -> str:
    pattern = r"/s/([a-zA-Z0-9]+)"
    match = re.search(pattern, url)
    if match:
        return match.group(1)
    return ""


def get_app_dir():
    if getattr(sys, 'frozen', False):
        return os.path.dirname(sys.executable)
    else:
        return os.path.dirname(os.path.abspath(__file__))


def parse_baidu_links(text):
    matches1 = re.finditer(r'(https?://)?pan\.baidu\.com/s/(1[a-zA-Z0-9\-_]+)', text)
    matches2 = re.finditer(r'(https?://)?pan\.baidu\.com/init\?surl=([a-zA-Z0-9\-_]+)', text)
    
    links = []
    
    def find_pwd_nearby(start_pos):
        sub_text = text[start_pos:start_pos+100]
        query_match = re.match(r'^[^\s?]*\?[^\s]*pwd=([a-zA-Z0-9]{4})', sub_text)
        if query_match:
            return query_match.group(1)
            
        label_match = re.search(r'(提取码|密码|提取密码|码|pwd|验码|验证码|提取|密)[:：\s]*([a-zA-Z0-9]{4})', sub_text, re.IGNORECASE)
        if label_match:
            return label_match.group(2)
            
        word_match = re.search(r'^\s+([a-zA-Z0-9]{4})\b', sub_text)
        if word_match:
            return word_match.group(1)
            
        return ""

    for m in matches1:
        full_url = m.group(0)
        key = m.group(2)
        pwd = find_pwd_nearby(m.end())
        url_pwd_match = re.search(r'[?&]pwd=([a-zA-Z0-9]{4})', full_url)
        if url_pwd_match:
            pwd = url_pwd_match.group(1)
        links.append({
            "raw_url": full_url,
            "key": key,
            "surl": key[1:] if key.startswith("1") else key,
            "pwd": pwd
        })
        
    for m in matches2:
        full_url = m.group(0)
        surl = m.group(2)
        key = "1" + surl
        pwd = find_pwd_nearby(m.end())
        url_pwd_match = re.search(r'[?&]pwd=([a-zA-Z0-9]{4})', full_url)
        if url_pwd_match:
            pwd = url_pwd_match.group(1)
        links.append({
            "raw_url": full_url,
            "key": key,
            "surl": surl,
            "pwd": pwd
        })
        
    seen = set()
    unique_links = []
    for item in links:
        if item["key"] not in seen:
            seen.add(item["key"])
            unique_links.append(item)
    return unique_links


class BaiduBatchTransfer:
    BASE_URL = "https://pan.baidu.com"

    def __init__(self, cookie: str):
        self.session = requests.Session()
        self.session.headers.update({
            "User-Agent": "Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/110.0.0.0 Safari/537.36",
            "Referer": "https://pan.baidu.com/disk/home"
        })
        self.set_cookie(cookie)
        self.bdstoken = None
        self.get_bdstoken()

    def set_cookie(self, cookie_str: str):
        cookie_str = cookie_str.strip()
        if "=" not in cookie_str:
            self.session.cookies.set("BDUSS", cookie_str, domain=".baidu.com")
            return
        for part in cookie_str.split(";"):
            part = part.strip()
            if not part:
                continue
            if "=" in part:
                k, v = part.split("=", 1)
                self.session.cookies.set(k, v, domain=".baidu.com")

    def get_bdstoken(self):
        url = f"{self.BASE_URL}/api/gettemplatevars"
        params = {
            "fields": '["bdstoken"]',
            "clienttype": "0",
            "app_id": "250528",
            "web": "1"
        }
        try:
            resp = self.session.get(url, params=params, timeout=10)
            if resp.status_code == 200:
                res_json = resp.json()
                if res_json.get("errno") == 0 and "result" in res_json:
                    self.bdstoken = res_json["result"].get("bdstoken")
                    if self.bdstoken:
                        return
        except Exception as e:
            logging.error(f"get_bdstoken API 获取失败: {e}")

        try:
            resp = self.session.get(f"{self.BASE_URL}/disk/home", timeout=10)
            if resp.status_code == 200:
                match = re.search(r'bdstoken["\']\s*[:=]\s*["\'](\w+)["\']', resp.text)
                if match:
                    self.bdstoken = match.group(1)
        except Exception as e:
            logging.error(f"get_bdstoken 页面解析失败: {e}")

    def get_account_info(self):
        url = f"{self.BASE_URL}/api/uinfo"
        params = {
            "clienttype": "0",
            "app_id": "250528",
            "web": "1"
        }
        try:
            resp = self.session.get(url, params=params, timeout=10)
            if resp.status_code == 200:
                res_json = resp.json()
                if res_json.get("errno") == 0:
                    return res_json
        except Exception as e:
            logging.error(f"get_account_info API 获取失败: {e}")

        try:
            resp = self.session.get(f"{self.BASE_URL}/disk/home", timeout=10)
            if resp.status_code == 200:
                match = re.search(r'["\']username["\']\s*[:=]\s*["\']([^"\']+)["\']', resp.text)
                if match:
                    return {
                        "errno": 0,
                        "nick_name": match.group(1),
                        "baidu_name": match.group(1)
                    }
        except Exception as e:
            logging.error(f"get_account_info 页面解析失败: {e}")
        return None

    def verify_share(self, surl, pwd):
        url = f"{self.BASE_URL}/share/verify"
        params = {
            "surl": surl,
            "clienttype": "0",
            "app_id": "250528",
            "web": "1"
        }
        data = {
            "pwd": pwd,
            "vcode": "",
            "vcode_str": ""
        }
        headers = {"Referer": f"https://pan.baidu.com/s/1{surl}"}
        try:
            resp = self.session.post(url, params=params, data=data, headers=headers, timeout=10)
            if resp.status_code == 200:
                return resp.json()
        except Exception as e:
            logging.error(f"verify_share 请求失败: {e}")
        return None

    def get_share_list(self, surl, pwd):
        share_page_url = f"https://pan.baidu.com/s/1{surl}?pwd={pwd}"
        headers = {
            "User-Agent": "Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/110.0.0.0 Safari/537.36",
            "Referer": "https://pan.baidu.com/disk/home"
        }
        shareid = None
        uk = None
        try:
            resp_page = self.session.get(share_page_url, headers=headers, timeout=10)
            if resp_page.status_code == 200:
                shareid_match = re.search(r'shareid["\']?\s*[:=]\s*["\']?(\d+)["\']?', resp_page.text)
                uk_match = re.search(r'share_uk["\']?\s*[:=]\s*["\']?(\d+)["\']?', resp_page.text)
                if not uk_match:
                    uk_match = re.search(r'uk["\']?\s*[:=]\s*["\']?(\d+)["\']?', resp_page.text)
                if shareid_match and uk_match:
                    shareid = shareid_match.group(1)
                    uk = uk_match.group(1)
        except Exception as e:
            logging.error(f"get_share_list 页面解析失败: {e}")

        if not shareid or not uk:
            return None

        url = f"{self.BASE_URL}/share/list"
        params = {
            "shareid": shareid,
            "uk": uk,
            "root": "1",
            "page": "1",
            "num": "100",
            "dir": "/",
            "channel": "chunlei",
            "clienttype": "0",
            "web": "1",
            "app_id": "250528"
        }
        try:
            resp = self.session.get(url, params=params, headers={"Referer": share_page_url}, timeout=10)
            if resp.status_code == 200:
                res_json = resp.json()
                if res_json.get("errno") == 0:
                    res_json["shareid"] = shareid
                    res_json["uk"] = uk
                    return res_json
        except Exception as e:
            logging.error(f"get_share_list API 获取失败: {e}")
        return None

    def exists(self, path):
        path = path.rstrip('/')
        if not path or path == "/":
            return True
        import posixpath
        parent_dir = posixpath.dirname(path)
        name = posixpath.basename(path)
        
        items = self.list_directory(parent_dir)
        for item in items:
            if item.get("server_filename") == name:
                return True
        return False

    def mkdir(self, path):
        if self.exists(path):
            return {"errno": 0, "msg": "already exists"}
        url = f"{self.BASE_URL}/api/create"
        params = {
            "a": "commit",
            "bdstoken": self.bdstoken,
            "channel": "chunlei",
            "web": "1",
            "app_id": "250528",
            "clienttype": "0"
        }
        data = {
            "path": path,
            "isdir": "1",
            "block_list": "[]",
            "method": "post"
        }
        try:
            resp = self.session.post(url, params=params, data=data, timeout=10)
            if resp.status_code == 200:
                return resp.json()
        except Exception as e:
            logging.error(f"mkdir 请求失败: {e}")
        return None

    def transfer_share(self, shareid, uk, fsid_list, dest_path, sekey):
        import urllib.parse
        decoded_sekey = urllib.parse.unquote(sekey)
        url = f"{self.BASE_URL}/share/transfer"
        params = {
            "shareid": shareid,
            "from": uk,
            "ondup": "newcopy",
            "async": "1",
            "bdstoken": self.bdstoken,
            "channel": "chunlei",
            "web": "1",
            "app_id": "250528",
            "clienttype": "0"
        }
        data = {
            "fsidlist": json.dumps(fsid_list),
            "path": dest_path,
            "sekey": decoded_sekey
        }
        try:
            resp = self.session.post(url, params=params, data=data, timeout=10)
            if resp.status_code == 200:
                return resp.json()
        except Exception as e:
            logging.error(f"transfer_share 请求失败: {e}")
        return None

    def list_directory(self, path):
        url = f"{self.BASE_URL}/api/list"
        params = {
            "dir": path,
            "bdstoken": self.bdstoken,
            "channel": "chunlei",
            "web": "1",
            "app_id": "250528",
            "clienttype": "0",
            "limit": "1000"
        }
        try:
            resp = self.session.get(url, params=params, timeout=10)
            if resp.status_code == 200:
                res_json = resp.json()
                if res_json.get("errno") == 0:
                    return res_json.get("list", [])
        except Exception as e:
            logging.error(f"list_directory 请求失败: {e}")
        return []

    def rename(self, path, newname):
        url = f"{self.BASE_URL}/api/filemanager"
        params = {
            "opera": "rename",
            "bdstoken": self.bdstoken,
            "channel": "chunlei",
            "web": "1",
            "app_id": "250528",
            "clienttype": "0"
        }
        data = {
            "filelist": json.dumps([{"path": path, "newname": newname}])
        }
        try:
            resp = self.session.post(url, params=params, data=data, timeout=10)
            if resp.status_code == 200:
                return resp.json()
        except Exception as e:
            logging.error(f"rename 请求失败: {e}")
        return None

    def delete(self, path_list):
        url = f"{self.BASE_URL}/api/filemanager"
        params = {
            "opera": "delete",
            "bdstoken": self.bdstoken,
            "channel": "chunlei",
            "web": "1",
            "app_id": "250528",
            "clienttype": "0"
        }
        data = {
            "filelist": json.dumps(path_list)
        }
        try:
            resp = self.session.post(url, params=params, data=data, timeout=10)
            if resp.status_code == 200:
                return resp.json()
        except Exception as e:
            logging.error(f"delete 请求失败: {e}")
        return None

    def share_files(self, fid_list, pwd):
        url = f"{self.BASE_URL}/share/set"
        params = {
            "channel": "chunlei",
            "clienttype": "0",
            "web": "1",
            "app_id": "250528",
            "bdstoken": self.bdstoken
        }
        data = {
            "fid_list": json.dumps(fid_list),
            "pwd": pwd,
            "period": "0",
            "schannel": "4"
        }
        try:
            resp = self.session.post(url, params=params, data=data, timeout=10)
            if resp.status_code == 200:
                return resp.json()
        except Exception as e:
            logging.error(f"share_files 请求失败: {e}")
        return None


class QuarkGUITool:
    CONFIG_FILE = os.path.join(get_app_dir(), "gui_config.json")

    def __init__(self, root):
        self.root = root
        self.root.title("SmartQuark - 智能转存与广告清理助手 v1.0")
        self.root.geometry("980x780")
        self.root.minsize(920, 720)

        # 加载配置
        self.config = self.load_config()

        # UI 字体及颜色配置
        self.font_title = ("Microsoft YaHei", 10, "bold")
        self.font_normal = ("Microsoft YaHei", 10)
        self.bg_color = "#f4f5f7"
        
        self.root.configure(bg=self.bg_color)
        
        self.transfer_thread = None
        self.stop_requested = False
        self.processed_resources = []

        self.create_widgets()
        self.load_values_from_config()

    def create_widgets(self):
        # 1. 顶部配置框
        top_frame = tk.LabelFrame(self.root, text="基本参数配置", font=self.font_title, bg=self.bg_color, padx=10, pady=10)
        top_frame.pack(side=tk.TOP, fill=tk.X, padx=15, pady=10)

        # Quark Cookie 输入
        tk.Label(top_frame, text="夸克 Cookie:", font=self.font_normal, bg=self.bg_color).grid(row=0, column=0, sticky=tk.W, pady=5)
        self.cookie_entry = ttk.Entry(top_frame, font=self.font_normal, show="*")
        self.cookie_entry.grid(row=0, column=1, columnspan=3, sticky=tk.EW, padx=5, pady=5)
        top_frame.columnconfigure(1, weight=1)

        self.cookie_toggle_btn = ttk.Button(top_frame, text="显示", width=6, command=lambda: self.toggle_cookie_visibility(self.cookie_entry, self.cookie_toggle_btn))
        self.cookie_toggle_btn.grid(row=0, column=4, padx=5, pady=5, sticky=tk.EW)

        self.remember_var = tk.BooleanVar(value=True)
        self.remember_cb = ttk.Checkbutton(top_frame, text="记住 Cookie", variable=self.remember_var)
        self.remember_cb.grid(row=0, column=5, padx=5, pady=5)

        # Baidu Cookie 输入
        tk.Label(top_frame, text="百度 Cookie:", font=self.font_normal, bg=self.bg_color).grid(row=1, column=0, sticky=tk.W, pady=5)
        self.baidu_cookie_entry = ttk.Entry(top_frame, font=self.font_normal, show="*")
        self.baidu_cookie_entry.grid(row=1, column=1, columnspan=3, sticky=tk.EW, padx=5, pady=5)

        self.baidu_cookie_toggle_btn = ttk.Button(top_frame, text="显示", width=6, command=lambda: self.toggle_cookie_visibility(self.baidu_cookie_entry, self.baidu_cookie_toggle_btn))
        self.baidu_cookie_toggle_btn.grid(row=1, column=4, padx=5, pady=5, sticky=tk.EW)

        self.clear_cookie_btn = ttk.Button(top_frame, text="清空 Cookie", command=self.clear_cookies)
        self.clear_cookie_btn.grid(row=1, column=5, padx=5, pady=5, sticky=tk.EW)

        # 日期 & 敏感词输入
        tk.Label(top_frame, text="转存日期:", font=self.font_normal, bg=self.bg_color).grid(row=2, column=0, sticky=tk.W, pady=5)
        self.date_entry = ttk.Entry(top_frame, font=self.font_normal, width=15)
        self.date_entry.grid(row=2, column=1, sticky=tk.W, padx=5, pady=5)
        self.date_entry.insert(0, datetime.now().strftime("%Y/%m/%d"))
        
        tk.Label(top_frame, text=" (格式: YYYY/MM/DD) ", font=self.font_normal, bg=self.bg_color, fg="#666").grid(row=2, column=2, sticky=tk.W)

        tk.Label(top_frame, text="广告敏感词 (逗号分隔):", font=self.font_normal, bg=self.bg_color).grid(row=2, column=3, sticky=tk.E, padx=5, pady=5)
        self.kw_entry = ttk.Entry(top_frame, font=self.font_normal)
        self.kw_entry.grid(row=2, column=4, columnspan=2, sticky=tk.EW, padx=5, pady=5)
        self.kw_entry.insert(0, "微信,扫码,获取更多,更多资源,加群,进群,领取,公众号,防走失,防失联,加好友,福利,5分打印,必看,必读,加微,更多课程,免费大放送,更多整理")

        # 2. 底部操作栏
        btn_frame = tk.Frame(self.root, bg=self.bg_color)
        btn_frame.pack(side=tk.BOTTOM, fill=tk.X, padx=15, pady=(0, 15))

        self.start_btn = tk.Button(btn_frame, text="开始批量转存与广告清理", font=("Microsoft YaHei", 11, "bold"), bg="#28a745", fg="white", activebackground="#218838", activeforeground="white", relief=tk.FLAT, pady=8, command=self.start_transfer)
        self.start_btn.pack(side=tk.LEFT, fill=tk.X, expand=True, padx=(0, 5))

        self.stop_btn = tk.Button(btn_frame, text="停止任务", font=("Microsoft YaHei", 11, "bold"), bg="#dc3545", fg="white", activebackground="#c82333", activeforeground="white", relief=tk.FLAT, pady=8, state="disabled", width=15, command=self.stop_transfer)
        self.stop_btn.pack(side=tk.RIGHT, padx=(5, 0))

        # 3. 下部日志显示区
        bottom_frame = tk.Frame(self.root, bg=self.bg_color)
        bottom_frame.pack(side=tk.BOTTOM, fill=tk.X, padx=15, pady=10)
        bottom_frame.columnconfigure(0, weight=3)
        bottom_frame.columnconfigure(1, weight=2)

        # 执行日志
        log_frame = tk.LabelFrame(bottom_frame, text="运行状态日志", font=self.font_title, bg=self.bg_color, padx=5, pady=5)
        log_frame.grid(row=0, column=0, sticky=tk.NSEW, padx=(0, 5))
        self.log_text = tk.Text(log_frame, font=self.font_normal, height=8, bg="#1e1e1e", fg="#d4d4d4", state="disabled", wrap=tk.WORD)
        self.log_text.pack(fill=tk.BOTH, expand=True)

        # 广告清理与重命名日志
        ad_frame = tk.LabelFrame(bottom_frame, text="广告清理与重命名日志", font=self.font_title, bg=self.bg_color, padx=5, pady=5)
        ad_frame.grid(row=0, column=1, sticky=tk.NSEW, padx=(5, 0))
        self.ad_log_text = tk.Text(ad_frame, font=self.font_normal, height=8, bg="#1e1e1e", fg="#ff6b6b", state="disabled", wrap=tk.WORD)
        self.ad_log_text.pack(fill=tk.BOTH, expand=True)

        # 4. 中部文案处理区
        mid_frame = tk.Frame(self.root, bg=self.bg_color)
        mid_frame.pack(side=tk.TOP, fill=tk.BOTH, expand=True, padx=15, pady=5)
        mid_frame.columnconfigure(0, weight=1)
        mid_frame.columnconfigure(1, weight=1)
        mid_frame.rowconfigure(0, weight=1)

        # 左侧输入
        left_frame = tk.LabelFrame(mid_frame, text=" 1. 请粘贴包含旧夸克或百度网盘链接的文案 ", font=self.font_title, bg=self.bg_color, padx=5, pady=5)
        left_frame.grid(row=0, column=0, sticky=tk.NSEW, padx=(0, 5))
        
        self.clear_btn = ttk.Button(left_frame, text="一键清空", command=self.clear_input)
        self.clear_btn.pack(side=tk.BOTTOM, fill=tk.X, pady=(5, 0))
        
        self.input_text = tk.Text(left_frame, font=self.font_normal, wrap=tk.WORD, undo=True)
        self.input_text.pack(side=tk.TOP, fill=tk.BOTH, expand=True)

        # 右侧输出
        right_frame = tk.LabelFrame(mid_frame, text=" 2. 转换替换后的新文案 ", font=self.font_title, bg=self.bg_color, padx=5, pady=5)
        right_frame.grid(row=0, column=1, sticky=tk.NSEW, padx=(5, 0))

        right_btn_frame = tk.Frame(right_frame, bg=self.bg_color)
        right_btn_frame.pack(side=tk.BOTTOM, fill=tk.X, pady=(5, 0))

        self.copy_btn = ttk.Button(right_btn_frame, text="一键复制新文案", command=self.copy_output)
        self.copy_btn.pack(side=tk.LEFT, fill=tk.X, expand=True, padx=(0, 2))

        self.export_btn = ttk.Button(right_btn_frame, text="一键导出 Excel", state="disabled", command=self.export_to_template)
        self.export_btn.pack(side=tk.RIGHT, fill=tk.X, expand=True, padx=(2, 0))

        self.output_text = tk.Text(right_frame, font=self.font_normal, wrap=tk.WORD)
        self.output_text.pack(side=tk.TOP, fill=tk.BOTH, expand=True)

    def load_config(self):
        if os.path.exists(self.CONFIG_FILE):
            try:
                with open(self.CONFIG_FILE, "r", encoding="utf-8") as f:
                    return json.load(f)
            except Exception as e:
                logging.error(f"读取配置失败: {e}")
        return {}

    def save_config(self):
        config_data = {
            "cookie": self.cookie_entry.get().strip() if self.remember_var.get() else "",
            "baidu_cookie": self.baidu_cookie_entry.get().strip() if self.remember_var.get() else "",
            "remember": self.remember_var.get(),
            "keywords": self.kw_entry.get().strip(),
        }
        try:
            with open(self.CONFIG_FILE, "w", encoding="utf-8") as f:
                json.dump(config_data, f, ensure_ascii=False, indent=2)
        except Exception as e:
            logging.error(f"保存配置失败: {e}")

    def load_values_from_config(self):
        if self.config:
            if self.config.get("remember"):
                self.cookie_entry.insert(0, self.config.get("cookie", ""))
                self.baidu_cookie_entry.insert(0, self.config.get("baidu_cookie", ""))
                self.remember_var.set(True)
            else:
                self.remember_var.set(False)
            
            if self.config.get("keywords"):
                saved_kws = [k.strip() for k in self.config.get("keywords").split(",") if k.strip()]
                defaults = ["微信", "扫码", "获取更多", "更多资源", "加群", "进群", "领取", "公众号", "防走失", "防失联", "加好友", "福利", "5分打印", "必看", "必读", "加微", "更多课程", "免费大放送", "更多整理"]
                for d in defaults:
                    if d not in saved_kws:
                        saved_kws.append(d)
                self.kw_entry.delete(0, tk.END)
                self.kw_entry.insert(0, ",".join(saved_kws))

    def log_message(self, message):
        timestamp = datetime.now().strftime("%H:%M:%S")
        formatted = f"[{timestamp}] {message}\n"
        self.root.after(0, self._insert_log, formatted)

    def _insert_log(self, text):
        self.log_text.configure(state="normal")
        self.log_text.insert(tk.END, text)
        self.log_text.see(tk.END)
        self.log_text.configure(state="disabled")

    def log_ad_deletion(self, message):
        timestamp = datetime.now().strftime("%H:%M:%S")
        formatted = f"[{timestamp}] {message}\n"
        self.root.after(0, self._insert_ad_log, formatted)

    def _insert_ad_log(self, text):
        self.ad_log_text.configure(state="normal")
        self.ad_log_text.insert(tk.END, text)
        self.ad_log_text.see(tk.END)
        self.ad_log_text.configure(state="disabled")

    def update_output_text(self, text):
        self.output_text.delete("1.0", tk.END)
        self.output_text.insert("1.0", text)

    def toggle_cookie_visibility(self, entry, button):
        if entry.cget("show"):
            entry.configure(show="")
            button.configure(text="隐藏")
        else:
            entry.configure(show="*")
            button.configure(text="显示")

    def copy_output(self):
        content = self.output_text.get("1.0", tk.END).strip()
        if not content:
            messagebox.showwarning("警告", "没有可复制的内容！")
            return
        self.root.clipboard_clear()
        self.root.clipboard_append(content)
        messagebox.showinfo("成功", "文案已成功复制到剪贴板！")

    def export_to_template(self):
        if not self.processed_resources:
            messagebox.showwarning("警告", "当前没有已转换的资源数据可以导出！")
            return

        default_dir = get_app_dir()
        default_file = f"资源导入_{datetime.now().strftime('%Y%m%d_%H%M%S')}.xlsx"
        
        save_path = filedialog.asksaveasfilename(
            initialdir=default_dir,
            initialfile=default_file,
            defaultextension=".xlsx",
            filetypes=[("Excel files", "*.xlsx"), ("All files", "*.*")],
            title="选择保存导出文件的位置"
        )
        if not save_path:
            return

        try:
            wb = openpyxl.Workbook()
            ws = wb.active
            ws.title = "资源导入"
            ws.cell(row=1, column=1, value="资源名称")
            ws.cell(row=1, column=2, value="资源地址")
            
            row = 2
            for res_name, res_link in self.processed_resources:
                ws.cell(row=row, column=1, value=res_name)
                ws.cell(row=row, column=2, value=res_link)
                row += 1
                
            wb.save(save_path)
            wb.close()
            messagebox.showinfo("成功", f"成功导出 {len(self.processed_resources)} 条资源数据到 Excel！\n保存路径:\n{save_path}")
        except Exception as e:
            messagebox.showerror("错误", f"导出失败！错误信息:\n{str(e)}")

    def clear_input(self):
        self.input_text.delete("1.0", tk.END)

    def clear_cookies(self):
        self.cookie_entry.delete(0, tk.END)
        self.baidu_cookie_entry.delete(0, tk.END)
        self.save_config()
        messagebox.showinfo("成功", "Cookie 已清空并保存！")

    def start_transfer(self):
        cookie = self.cookie_entry.get().strip()
        baidu_cookie = self.baidu_cookie_entry.get().strip()
        
        input_content = self.input_text.get("1.0", tk.END).strip()
        if not input_content:
            messagebox.showerror("错误", "请输入或粘贴待转换的微信文案！")
            return

        has_quark = "pan.quark.cn" in input_content or "quark.cn/s/" in input_content
        has_baidu = "pan.baidu.com" in input_content

        if not has_quark and not has_baidu:
            messagebox.showerror("错误", "未在文案中检测到有效的夸克或百度网盘分享链接！")
            return

        if has_quark and not cookie:
            messagebox.showerror("错误", "检测到文案中包含夸克链接，请输入有效的夸克 Cookie！")
            return

        if has_baidu and not baidu_cookie:
            messagebox.showerror("错误", "检测到文案中包含百度网盘链接，请输入有效的百度 Cookie！")
            return

        date_str = self.date_entry.get().strip().replace("-", "/").replace(".", "/")
        if not re.match(r"^\d{4}/\d{2}/\d{2}$", date_str):
            messagebox.showerror("错误", "日期格式不正确！请输入形如 YYYY/MM/DD 的日期。")
            return

        # 临时保存设置
        self.save_config()

        # 初始化状态并启动后台线程
        self.processed_resources = []
        self.export_btn.configure(state="disabled")
        self.copy_btn.configure(state="disabled")
        self.clear_btn.configure(state="disabled")
        self.start_btn.configure(state="disabled", bg="#6c757d")
        self.stop_btn.configure(state="normal")
        self.stop_requested = False
        
        # 清空输出与日志
        self.update_output_text("")
        self.log_text.configure(state="normal")
        self.log_text.delete("1.0", tk.END)
        self.log_text.configure(state="disabled")
        self.ad_log_text.configure(state="normal")
        self.ad_log_text.delete("1.0", tk.END)
        self.ad_log_text.configure(state="disabled")

        kw_str = self.kw_entry.get().strip()
        self.transfer_thread = threading.Thread(target=self.run_transfer_process, args=(cookie, baidu_cookie, date_str, input_content, kw_str))
        self.transfer_thread.daemon = True
        self.transfer_thread.start()

    def stop_transfer(self):
        if self.transfer_thread and self.transfer_thread.is_alive():
            self.stop_requested = True
            self.log_message("收到停止指令，正在安全退出当前转存任务...")
            self.stop_btn.configure(state="disabled")

    def run_transfer_process(self, cookie, baidu_cookie, date_str, original_text, kw_str):
      try:
        has_quark = "pan.quark.cn" in original_text or "quark.cn/s/" in original_text
        has_baidu = "pan.baidu.com" in original_text

        blacklist_keywords = [k.strip() for k in re.split(r"[,，]", kw_str) if k.strip()]
        self.log_message(f"过滤广告敏感词: {blacklist_keywords}")

        dest_path = f"/微信群日更/{date_str}"

        quark_transfer = None
        if has_quark:
            self.log_message("正在验证 夸克 Cookie 有效性...")
            quark_transfer = QuarkBatchTransfer(cookie)
            account_info = quark_transfer.get_account_info()
            if not account_info:
                self.log_message("夸克 Cookie 验证失败！请检查您的 Cookie 是否已过期。")
                self.root.after(0, self.finish_process, "failed")
                return
            nickname = account_info.get("nickname", "未知用户")
            self.log_message(f"夸克登录成功！账号昵称: {nickname}")

        baidu_transfer = None
        if has_baidu:
            self.log_message("正在验证 百度网盘 Cookie 有效性...")
            baidu_transfer = BaiduBatchTransfer(baidu_cookie)
            account_info = baidu_transfer.get_account_info()
            if not account_info:
                self.log_message("百度网盘 Cookie 验证失败！请检查您的 Cookie 是否已过期。")
                self.root.after(0, self.finish_process, "failed")
                return
            nickname = account_info.get("nick_name") or account_info.get("baidu_name") or "未知用户"
            self.log_message(f"百度网盘登录成功！账号昵称: {nickname}")

        current_text = original_text

        # ==========================================
        # 1. 夸克网盘转存处理
        # ==========================================
        if has_quark and quark_transfer:
            self.log_message("==== 开始处理夸克网盘任务 ====")
            self.log_message(f"夸克目标网盘目录: {dest_path}")
            
            fids_info = quark_transfer.get_fids([dest_path])
            dest_fid = None
            for item in fids_info:
                if item.get("file_path") == dest_path:
                    dest_fid = item.get("fid")
                    break
            
            if self.stop_requested:
                self.log_message("任务已中止。")
                self.root.after(0, self.finish_process, "stopped")
                return

            if not dest_fid:
                self.log_message("夸克目标日期目录不存在，正在自动创建...")
                mkdir_res = quark_transfer.mkdir(dest_path)
                if mkdir_res:
                    dest_fid = mkdir_res.get("fid")
                    self.log_message(f"日期目录创建成功，fid: {dest_fid}")
                else:
                    self.log_message("创建日期目录失败，夸克任务跳过。")
                    dest_fid = None
            else:
                self.log_message(f"目录已存在，使用当前 fid: {dest_fid}")

            if dest_fid:
                quark_links = list(dict.fromkeys(re.findall(r"https://pan\.quark\.cn/s/[a-zA-Z0-9]+", current_text)))
                total_quark_links = len(quark_links)
                self.log_message(f"在文案中共识别出 {total_quark_links} 个独立夸克网盘链接。")

                batch_fid = None
                for idx, old_link in enumerate(quark_links, 1):
                    if self.stop_requested:
                        self.log_message("任务被用户中止。")
                        break

                    self.log_message(f"--------------------------------------------------")
                    self.log_message(f"Quark [{idx}/{total_quark_links}] 正在处理链接: {old_link}")
                    
                    pwd_id = get_id_from_url(old_link)
                    if not pwd_id:
                        self.log_message("解析链接 ID 失败，跳过。")
                        continue

                    stoken = quark_transfer.get_stoken(pwd_id)
                    if not stoken:
                        self.log_message("获取分享 Token 失败，链接可能已失效或被风控限制。")
                        continue

                    detail_list = quark_transfer.get_detail(pwd_id, stoken)
                    if not detail_list:
                        self.log_message("获取分享详情失败，或者分享文件夹已空。")
                        continue

                    book_name = detail_list[0].get("file_name", "未知资源")
                    self.log_message(f"识别出资源名称: {book_name}")

                    if batch_fid is None:
                        safe_book_name = re.sub(r'[\\/:*?"<>|]', "_", book_name)
                        batch_folder_name = f"{safe_book_name}等"
                        batch_folder_path = f"{dest_path}/{batch_folder_name}"
                        self.log_message(f"创建夸克批次归档文件夹: {batch_folder_path}")
                        
                        fids_info = quark_transfer.get_fids([batch_folder_path])
                        for item in fids_info:
                            if item.get("file_path") == batch_folder_path:
                                batch_fid = item.get("fid")
                                break
                        
                        if not batch_fid:
                            mkdir_res = quark_transfer.mkdir(batch_folder_path)
                            if mkdir_res:
                                batch_fid = mkdir_res.get("fid")
                                self.log_message(f"批次归档文件夹创建成功，fid: {batch_fid}")
                            else:
                                self.log_message("创建批次归档文件夹失败，将直接保存至日期根目录。")
                                batch_fid = dest_fid

                    target_save_fid = batch_fid if batch_fid else dest_fid
                    fid_list = [item["fid"] for item in detail_list]
                    fid_token_list = [item["share_fid_token"] for item in detail_list]

                    save_task_id = quark_transfer.save_file(fid_list, fid_token_list, target_save_fid, pwd_id, stoken)
                    if not save_task_id:
                        self.log_message("创建转存任务失败。")
                        continue

                    save_res = quark_transfer.query_task(save_task_id, stop_flag=lambda: self.stop_requested)
                    if not save_res:
                        self.log_message("转存任务在网盘执行失败。")
                        continue

                    save_as_top_fids = save_res.get("data", {}).get("save_as", {}).get("save_as_top_fids")
                    if not save_as_top_fids:
                        self.log_message("未取得保存后的文件 ID 列表。")
                        continue

                    for saved_fid in save_as_top_fids:
                        clean_ads_recursively(quark_transfer, saved_fid, book_name, blacklist_keywords, self.log_ad_deletion, stop_flag=lambda: self.stop_requested)
                    if self.stop_requested:
                        self.log_message("任务被用户中止。")
                        break

                    share_task_id = quark_transfer.share_files(save_as_top_fids, book_name)
                    if not share_task_id:
                        self.log_message("创建分享任务失败。")
                        continue

                    share_res = quark_transfer.query_task(share_task_id, stop_flag=lambda: self.stop_requested)
                    if not share_res:
                        self.log_message("重新分享任务执行失败。")
                        continue

                    share_id = share_res.get("data", {}).get("share_id")
                    if not share_id:
                        self.log_message("未获取到有效的分享 ID。")
                        continue

                    new_share_link = quark_transfer.get_share_link(share_id)
                    if not new_share_link:
                        self.log_message("获取新分享链接失败。")
                        continue

                    self.processed_resources.append((book_name, new_share_link))
                    self.log_message(f"转存及清理成功！新链接: {new_share_link}")

                    current_text = current_text.replace(old_link, new_share_link)
                    self.root.after(0, self.update_output_text, current_text)

                    if idx < total_quark_links or has_baidu:
                        sleep_time = random.uniform(3.0, 5.0)
                        self.log_message(f"休眠 {sleep_time:.2f} 秒以保护账号...")
                        if not sleep_with_stop(sleep_time, stop_flag=lambda: self.stop_requested):
                            self.log_message("任务被用户中止。")
                            break

        # ==========================================
        # 2. 百度网盘转存处理
        # ==========================================
        if has_baidu and baidu_transfer:
            self.log_message("\n==== 开始处理百度网盘任务 ====")
            self.log_message(f"百度网盘目标目录: {dest_path}")
            
            self.log_message("正在检查/创建百度网盘日期目录...")
            baidu_transfer.mkdir(dest_path)
            
            baidu_links = parse_baidu_links(current_text)
            total_baidu_links = len(baidu_links)
            self.log_message(f"在文案中共识别出 {total_baidu_links} 个独立百度网盘链接。")

            baidu_batch_path = None
            for idx, item in enumerate(baidu_links, 1):
                if self.stop_requested:
                    self.log_message("任务被用户中止。")
                    break

                raw_url = item["raw_url"]
                surl = item["surl"]
                pwd = item["pwd"]

                self.log_message(f"--------------------------------------------------")
                self.log_message(f"Baidu [{idx}/{total_baidu_links}] 正在处理链接: {raw_url} 密码: {pwd or '无'}")

                sekey = ""
                if pwd:
                    self.log_message("正在提交提取码验证...")
                    verify_res = baidu_transfer.verify_share(surl, pwd)
                    if not verify_res or verify_res.get("errno") != 0:
                        self.log_message(f"提取码验证失败！接口响应: {verify_res}")
                        continue
                    sekey = verify_res.get("randsk", "")
                    self.log_message("提取码验证成功！")

                self.log_message("正在获取分享资源详情...")
                share_list_res = baidu_transfer.get_share_list(surl, pwd)
                if not share_list_res or share_list_res.get("errno") != 0:
                    self.log_message(f"获取分享资源详情失败！接口响应: {share_list_res}")
                    continue

                list_data = share_list_res.get("list", [])
                if not list_data:
                    self.log_message("该分享链接中的文件列表为空，跳过。")
                    continue

                uk = share_list_res.get("uk")
                shareid = share_list_res.get("shareid")
                book_name = list_data[0].get("server_filename", "未知资源")
                self.log_message(f"识别出资源名称: {book_name}")

                if baidu_batch_path is None:
                    safe_book_name = re.sub(r'[\\/:*?"<>|]', "_", book_name)
                    batch_folder_name = f"{safe_book_name}等"
                    baidu_batch_path = f"{dest_path}/{batch_folder_name}"
                    self.log_message(f"创建百度批次归档文件夹: {baidu_batch_path}")
                    baidu_transfer.mkdir(baidu_batch_path)

                existing_items = baidu_transfer.list_directory(baidu_batch_path)
                existing_paths = {itm["path"] for itm in existing_items}

                fsid_list = [itm["fs_id"] for itm in list_data]
                self.log_message("正在发送转存请求...")
                transfer_res = baidu_transfer.transfer_share(shareid, uk, fsid_list, baidu_batch_path, sekey)
                if not transfer_res or transfer_res.get("errno") != 0:
                    self.log_message(f"转存失败！错误信息: {transfer_res}")
                    continue

                self.log_message("转存任务创建成功，等待同步数据...")
                if not sleep_with_stop(2, stop_flag=lambda: self.stop_requested):
                    self.log_message("任务被用户中止。")
                    break

                current_items = baidu_transfer.list_directory(baidu_batch_path)
                new_items = [itm for itm in current_items if itm["path"] not in existing_paths]
                if not new_items:
                    self.log_message("未能捕获到新转存的资源对象。")
                    continue

                self.log_message("正在递归清理新文件中的广告敏感词...")
                for itm in new_items:
                    clean_baidu_ads_recursively(baidu_transfer, itm["path"], book_name, blacklist_keywords, self.log_ad_deletion, stop_flag=lambda: self.stop_requested)
                if self.stop_requested:
                    self.log_message("任务被用户中止。")
                    break

                final_items = baidu_transfer.list_directory(baidu_batch_path)
                final_fids = {itm["fs_id"] for itm in final_items}
                
                fids_to_share = [itm["fs_id"] for itm in new_items if itm["fs_id"] in final_fids]
                if not fids_to_share:
                    self.log_message("清洗后无有效文件，无法重新分享。")
                    continue

                new_pwd = "".join(random.choices("abcdefghijklmnopqrstuvwxyz0123456789", k=4))
                self.log_message(f"正在创建新的分享链接，提取码: {new_pwd}")
                
                share_res = baidu_transfer.share_files(fids_to_share, new_pwd)
                if not share_res or share_res.get("errno") != 0:
                    self.log_message(f"创建分享链接失败：{share_res}")
                    continue

                new_link = share_res.get("shorturl") or share_res.get("link")
                if not new_link:
                    self.log_message("未获取到有效的分享 URL。")
                    continue

                if "?" in new_link:
                    export_link = f"{new_link}&pwd={new_pwd}"
                else:
                    export_link = f"{new_link}?pwd={new_pwd}"
                self.processed_resources.append((book_name, export_link))
                self.log_message(f"转存及清理成功！新分享链接: {new_link}")

                if pwd:
                    # 1. 先替换 URL 中的密码参数（如 URL 包含 ?pwd=旧密码）
                    url_pwd_pattern = re.escape(raw_url) + r'([?&]pwd=)' + re.escape(pwd)
                    current_text, url_pwd_count = re.subn(url_pwd_pattern, lambda m: new_link + m.group(1) + new_pwd, current_text)
                    
                    # 2. 再替换链接后面紧随的提取码文本。允许匹配已经更新为新密码的 URL 参数，或者旧密码参数，或者无参数
                    pattern = re.escape(raw_url) + r'((?:[?&]pwd=(?:' + re.escape(pwd) + r'|' + re.escape(new_pwd) + r'))?\s*(?:[^\w\n]*?(?:提取码|密码|提取密码|码|pwd|密)[:：\s=]*?)\s*)' + re.escape(pwd)
                    def repl(m):
                        return new_link + m.group(1) + new_pwd
                    current_text, count = re.subn(pattern, repl, current_text)
                    
                    if count == 0 and url_pwd_count == 0:
                        current_text = current_text.replace(raw_url, f"{new_link} 提取码: {new_pwd}")
                else:
                    current_text = current_text.replace(raw_url, f"{new_link} 提取码: {new_pwd}")

                self.root.after(0, self.update_output_text, current_text)

                if idx < total_baidu_links:
                    sleep_time = random.uniform(3.0, 5.0)
                    self.log_message(f"休眠 {sleep_time:.2f} 秒以保护账号...")
                    if not sleep_with_stop(sleep_time, stop_flag=lambda: self.stop_requested):
                        self.log_message("任务被用户中止。")
                        break

        self.log_message(f"\n==================================================")
        self.log_message(f"批量转存与广告清理流程结束！")
        self.log_message(f"==================================================")
        
        self.root.after(0, self.finish_process, "completed")

      except Exception as e:
        logging.error(f"run_transfer_process 未捕获异常: {e}", exc_info=True)
        self.log_message(f"发生未预期的错误: {e}")
        self.root.after(0, self.finish_process, "failed")

    def finish_process(self, status="completed"):
        self.start_btn.configure(state="normal", bg="#28a745")
        self.stop_btn.configure(state="disabled")
        self.copy_btn.configure(state="normal")
        self.clear_btn.configure(state="normal")
        if self.processed_resources:
            self.export_btn.configure(state="normal")
            
        if status == "stopped" or self.stop_requested:
            messagebox.showinfo("中止", "转存任务已手动中止！已成功替换部分链接。")
        elif status == "failed":
            messagebox.showerror("错误", "转存任务由于验证错误或其它异常已终止，请检查日志。")
        else:
            messagebox.showinfo("完成", "批量转存和链接替换已完成！")


if __name__ == '__main__':
    root = tk.Tk()
    app = QuarkGUITool(root)
    root.mainloop()
