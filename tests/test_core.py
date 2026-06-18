# -*- coding: utf-8 -*-
"""Comprehensive unit tests for SmartQuark core behavior.

The suite stays Python 3.8.10 compatible; no 3.9+ features are used.
"""

import sys
import os
import tempfile
import unittest
from unittest.mock import patch

# Ensure the project root is on the path so we can import the module.
sys.path.insert(0, os.path.join(os.path.dirname(__file__), ".."))

from smart_quark import clean_filename, parse_baidu_links, BaiduBatchTransfer
from smart_quark import classify_file_action
from smart_quark import sleep_with_stop
from smart_quark import QuarkGUITool


# ── clean_filename ──────────────────────────────────────────────────────────


class TestCleanFilenameBasic(unittest.TestCase):
    """Baseline / happy-path cases."""

    def test_no_keyword_match_returns_original(self):
        self.assertEqual(clean_filename("good_video.mp4", ["公众号"]), "good_video.mp4")

    def test_empty_keywords_returns_original(self):
        self.assertEqual(clean_filename("hello.pdf", []), "hello.pdf")

    def test_empty_keyword_string_skipped(self):
        # An empty-string keyword should be silently skipped.
        self.assertEqual(clean_filename("hello.pdf", [""]), "hello.pdf")

    def test_extension_preserved_after_cleaning(self):
        result = clean_filename("资料【公众号xxx】.pdf", ["公众号"])
        self.assertTrue(result.endswith(".pdf"), result)

    def test_various_extensions_preserved(self):
        for ext in [".mp4", ".zip", ".rar", ".epub", ".txt"]:
            result = clean_filename("资料【公众号xxx】" + ext, ["公众号"])
            self.assertTrue(result.endswith(ext), result)

    def test_base_empty_returns_empty_string(self):
        # If everything is stripped the function must return "".
        result = clean_filename("【公众号xxx】.jpg", ["公众号"])
        self.assertEqual(result, "")


class TestCleanFilenameBracketKeywords(unittest.TestCase):
    """Keyword wrapped in every supported bracket type."""

    def test_chinese_brackets(self):
        self.assertEqual(
            clean_filename("资料【公众号xxx】课件.pdf", ["公众号"]),
            "资料课件.pdf",
        )

    def test_square_brackets(self):
        self.assertEqual(
            clean_filename("资料[公众号xxx]课件.pdf", ["公众号"]),
            "资料课件.pdf",
        )

    def test_parentheses(self):
        self.assertEqual(
            clean_filename("资料(公众号xxx)课件.pdf", ["公众号"]),
            "资料课件.pdf",
        )

    def test_chinese_parentheses(self):
        self.assertEqual(
            clean_filename("资料（公众号xxx）课件.pdf", ["公众号"]),
            "资料课件.pdf",
        )

    def test_curly_braces(self):
        self.assertEqual(
            clean_filename("资料{公众号xxx}课件.pdf", ["公众号"]),
            "资料课件.pdf",
        )


class TestCleanFilenameDomainRemoval(unittest.TestCase):
    """Domain / URL pattern detection and removal."""

    def test_domain_inside_chinese_brackets(self):
        self.assertEqual(
            clean_filename("file【免费分享：cunlove.cn】.pdf", []),
            "file.pdf",
        )

    def test_domain_inside_square_brackets(self):
        self.assertEqual(
            clean_filename("file[www.example.com].pdf", []),
            "file.pdf",
        )

    def test_domain_without_brackets_dash(self):
        result = clean_filename("file-cunlove.cn.pdf", [])
        self.assertEqual(result, "file.pdf")

    def test_domain_without_brackets_spaces(self):
        result = clean_filename("file cunlove.cn.pdf", [])
        self.assertEqual(result, "file.pdf")

    def test_domain_various_tlds(self):
        for domain in ["abc.net", "xyz.org", "test.vip", "demo.top", "hi.cc"]:
            result = clean_filename("data【{}】notes.pdf".format(domain), [])
            self.assertEqual(result, "datanotes.pdf", "Failed for domain: " + domain)

    def test_domain_combined_with_keyword(self):
        result = clean_filename("file【cunlove.cn】-公众号xxx.mp4", ["公众号"])
        self.assertEqual(result, "file.mp4")


class TestCleanFilenameSeparatorPrefix(unittest.TestCase):
    """Keyword preceded by a separator character (-, _, space, etc.)."""

    def test_dash_prefix(self):
        self.assertEqual(
            clean_filename("file-公众号xxx.mp4", ["公众号"]),
            "file.mp4",
        )

    def test_underscore_prefix(self):
        self.assertEqual(
            clean_filename("file_公众号xxx.mp4", ["公众号"]),
            "file.mp4",
        )

    def test_space_prefix(self):
        self.assertEqual(
            clean_filename("file 公众号xxx.mp4", ["公众号"]),
            "file.mp4",
        )

    def test_pipe_prefix(self):
        result = clean_filename("file|公众号xxx.mp4", ["公众号"])
        self.assertEqual(result, "file.mp4")


class TestCleanFilenameColonPattern(unittest.TestCase):
    """Keyword followed by a colon (Chinese or English) and arbitrary text."""

    def test_chinese_colon(self):
        # The colon pattern only strips keyword + colon + whitespace,
        # not the arbitrary text ("xxx") that follows.
        result = clean_filename("公众号：xxx resource.pdf", ["公众号"])
        self.assertEqual(result, "xxx resource.pdf")

    def test_english_colon(self):
        result = clean_filename("公众号:xxx resource.pdf", ["公众号"])
        self.assertEqual(result, "xxx resource.pdf")

    def test_colon_followed_by_space_only(self):
        # When keyword+colon is followed by nothing meaningful
        result = clean_filename("公众号： resource.pdf", ["公众号"])
        self.assertEqual(result, "resource.pdf")


class TestCleanFilenameMultipleKeywords(unittest.TestCase):
    """More than one keyword applied in sequence."""

    def test_two_keywords(self):
        result = clean_filename("资料【公众号xxx】-微信yyy.mp4", ["公众号", "微信"])
        self.assertEqual(result, "资料.mp4")

    def test_keywords_with_no_match_mixed(self):
        result = clean_filename("资料【公众号xxx】.mp4", ["公众号", "不存在"])
        self.assertEqual(result, "资料.mp4")


class TestCleanFilenameEdgeCases(unittest.TestCase):
    """Various edge / corner cases."""

    def test_no_extension(self):
        result = clean_filename("readme", ["公众号"])
        self.assertEqual(result, "readme")

    def test_filename_is_only_extension(self):
        # e.g. ".gitignore" – base_name is '' after ext split via os.path.splitext
        result = clean_filename(".gitignore", [])
        # os.path.splitext(".gitignore") => ('.gitignore', '')
        self.assertEqual(result, ".gitignore")

    def test_leading_trailing_separators_cleaned(self):
        # After keyword removal, leading/trailing separators are trimmed.
        result = clean_filename("-公众号xxx-资料.pdf", ["公众号"])
        self.assertEqual(result, "资料.pdf")


# ── parse_baidu_links ───────────────────────────────────────────────────────


class TestParseBaiduLinksBasic(unittest.TestCase):
    """Basic link + password extraction."""

    def test_single_link_with_tiquma_chinese_colon(self):
        text = "https://pan.baidu.com/s/1aBcDeFgH 提取码：ab12"
        result = parse_baidu_links(text)
        self.assertEqual(len(result), 1)
        self.assertEqual(result[0]["key"], "1aBcDeFgH")
        self.assertEqual(result[0]["pwd"], "ab12")

    def test_single_link_with_tiquma_english_colon(self):
        text = "https://pan.baidu.com/s/1aBcDeFgH 提取码:ab12"
        result = parse_baidu_links(text)
        self.assertEqual(len(result), 1)
        self.assertEqual(result[0]["pwd"], "ab12")

    def test_pwd_in_url_query(self):
        text = "https://pan.baidu.com/s/1aBcDeFgH?pwd=xy99"
        result = parse_baidu_links(text)
        self.assertEqual(len(result), 1)
        self.assertEqual(result[0]["pwd"], "xy99")

    def test_mima_label(self):
        text = "https://pan.baidu.com/s/1aBcDeFgH 密码：ab12"
        result = parse_baidu_links(text)
        self.assertEqual(len(result), 1)
        self.assertEqual(result[0]["pwd"], "ab12")

    def test_mima_english_colon(self):
        text = "https://pan.baidu.com/s/1aBcDeFgH 密码:ab12"
        result = parse_baidu_links(text)
        self.assertEqual(len(result), 1)
        self.assertEqual(result[0]["pwd"], "ab12")

    def test_pwd_label(self):
        text = "https://pan.baidu.com/s/1aBcDeFgH pwd:ab12"
        result = parse_baidu_links(text)
        self.assertEqual(len(result), 1)
        self.assertEqual(result[0]["pwd"], "ab12")


class TestParseBaiduLinksInitSurl(unittest.TestCase):
    """Links using the init?surl= format."""

    def test_surl_format(self):
        text = "https://pan.baidu.com/init?surl=AbCdEf 提取码：x1y2"
        result = parse_baidu_links(text)
        self.assertEqual(len(result), 1)
        self.assertEqual(result[0]["surl"], "AbCdEf")
        self.assertEqual(result[0]["key"], "1AbCdEf")
        self.assertEqual(result[0]["pwd"], "x1y2")

    def test_surl_format_without_password(self):
        text = "https://pan.baidu.com/init?surl=AbCdEf"
        result = parse_baidu_links(text)
        self.assertEqual(len(result), 1)
        self.assertEqual(result[0]["pwd"], "")


class TestParseBaiduLinksNoPwd(unittest.TestCase):
    """Link without any discoverable password."""

    def test_no_password(self):
        text = "https://pan.baidu.com/s/1aBcDeFgH"
        result = parse_baidu_links(text)
        self.assertEqual(len(result), 1)
        self.assertEqual(result[0]["pwd"], "")


class TestParseBaiduLinksMultiple(unittest.TestCase):
    """Multiple links in the same text block."""

    def test_two_different_links(self):
        text = (
            "https://pan.baidu.com/s/1aaaaBBBB 提取码：aa11\n"
            "https://pan.baidu.com/s/1ccccDDDD 提取码：bb22"
        )
        result = parse_baidu_links(text)
        self.assertEqual(len(result), 2)
        keys = {r["key"] for r in result}
        self.assertIn("1aaaaBBBB", keys)
        self.assertIn("1ccccDDDD", keys)

    def test_mixed_formats(self):
        text = (
            "https://pan.baidu.com/s/1aaaaBBBB?pwd=aa11\n"
            "https://pan.baidu.com/init?surl=ccccDDDD 密码:bb22"
        )
        result = parse_baidu_links(text)
        self.assertEqual(len(result), 2)


class TestParseBaiduLinksDedup(unittest.TestCase):
    """Duplicate links (same key) must be deduplicated."""

    def test_exact_duplicate(self):
        text = (
            "https://pan.baidu.com/s/1aBcDeFgH 提取码：ab12\n"
            "https://pan.baidu.com/s/1aBcDeFgH 提取码：ab12"
        )
        result = parse_baidu_links(text)
        self.assertEqual(len(result), 1)

    def test_duplicate_different_pwd_keeps_first(self):
        text = (
            "https://pan.baidu.com/s/1aBcDeFgH 提取码：ab12\n"
            "https://pan.baidu.com/s/1aBcDeFgH 提取码：zz99"
        )
        result = parse_baidu_links(text)
        self.assertEqual(len(result), 1)
        # First occurrence wins.
        self.assertEqual(result[0]["pwd"], "ab12")


class TestParseBaiduLinksNonBaiduIgnored(unittest.TestCase):
    """Only Baidu links are extracted; other vendors are ignored."""

    def test_quark_link_ignored(self):
        text = "https://pan.quark.cn/s/abc123 提取码：ab12"
        result = parse_baidu_links(text)
        self.assertEqual(len(result), 0)

    def test_mixed_quark_and_baidu(self):
        text = (
            "https://pan.quark.cn/s/abc123 提取码：ab12\n"
            "https://pan.baidu.com/s/1aBcDeFgH 提取码：cd34"
        )
        result = parse_baidu_links(text)
        self.assertEqual(len(result), 1)
        self.assertEqual(result[0]["key"], "1aBcDeFgH")


class TestParseBaiduLinksFields(unittest.TestCase):
    """Verify all returned dict fields."""

    def test_standard_link_fields(self):
        text = "https://pan.baidu.com/s/1aBcDeFgH?pwd=ab12"
        result = parse_baidu_links(text)
        item = result[0]
        self.assertIn("raw_url", item)
        self.assertIn("key", item)
        self.assertIn("surl", item)
        self.assertIn("pwd", item)
        self.assertEqual(item["key"], "1aBcDeFgH")
        self.assertEqual(item["surl"], "aBcDeFgH")  # key[1:]
        self.assertEqual(item["pwd"], "ab12")

    def test_surl_format_fields(self):
        text = "https://pan.baidu.com/init?surl=XyZ123 提取码：ab12"
        result = parse_baidu_links(text)
        item = result[0]
        self.assertEqual(item["surl"], "XyZ123")
        self.assertEqual(item["key"], "1XyZ123")
        self.assertEqual(item["pwd"], "ab12")


class TestParseBaiduLinksStandaloneCode(unittest.TestCase):
    """Password as a standalone 4-char code after the link."""

    def test_standalone_code_after_whitespace(self):
        text = "https://pan.baidu.com/s/1aBcDeFgH ab12"
        result = parse_baidu_links(text)
        self.assertEqual(len(result), 1)
        self.assertEqual(result[0]["pwd"], "ab12")


class TestParseBaiduLinksWithoutScheme(unittest.TestCase):
    """Links without http:// or https:// prefix."""

    def test_no_scheme(self):
        text = "pan.baidu.com/s/1aBcDeFgH 提取码：ab12"
        result = parse_baidu_links(text)
        self.assertEqual(len(result), 1)
        self.assertEqual(result[0]["pwd"], "ab12")


class TestParseBaiduLinksEmpty(unittest.TestCase):
    """Empty or irrelevant input."""

    def test_empty_string(self):
        self.assertEqual(parse_baidu_links(""), [])

    def test_no_links(self):
        self.assertEqual(parse_baidu_links("nothing special here"), [])


class TestExportToExcel(unittest.TestCase):
    """Excel export behavior."""

    def test_export_creates_blank_workbook_without_loading_legacy_template(self):
        with tempfile.TemporaryDirectory() as tmpdir:
            template_dir = os.path.join(tmpdir, "Import_Template")
            os.makedirs(template_dir)
            template_path = os.path.join(template_dir, "资源导入模板.xlsx")
            with open(template_path, "wb") as f:
                f.write(b"legacy template should not be read")

            save_path = os.path.join(tmpdir, "export.xlsx")
            gui = type("DummyGui", (), {})()
            gui.processed_resources = [("资源A", "https://pan.baidu.com/s/abc?pwd=1234")]

            with patch("smart_quark.get_app_dir", return_value=tmpdir), \
                 patch("smart_quark.filedialog.asksaveasfilename", return_value=save_path), \
                 patch("smart_quark.openpyxl.load_workbook", side_effect=AssertionError("legacy template was loaded")), \
                 patch("smart_quark.messagebox.showinfo") as showinfo, \
                 patch("smart_quark.messagebox.showerror") as showerror:
                QuarkGUITool.export_to_template(gui)

            showerror.assert_not_called()
            showinfo.assert_called_once()

            import openpyxl

            wb = openpyxl.load_workbook(save_path)
            ws = wb.active
            self.assertEqual(ws.cell(row=1, column=1).value, "资源名称")
            self.assertEqual(ws.cell(row=1, column=2).value, "资源地址")
            self.assertEqual(ws.cell(row=2, column=1).value, "资源A")
            self.assertEqual(ws.cell(row=2, column=2).value, "https://pan.baidu.com/s/abc?pwd=1234")
            wb.close()


class FakeEntry(object):
    def __init__(self, show="*"):
        self.options = {"show": show}

    def cget(self, key):
        return self.options.get(key)

    def configure(self, **kwargs):
        self.options.update(kwargs)


class FakeButton(object):
    def __init__(self):
        self.options = {}

    def configure(self, **kwargs):
        self.options.update(kwargs)


class TestCookieVisibility(unittest.TestCase):
    """Cookie input visibility toggling."""

    def test_toggle_cookie_visibility_shows_then_hides_value(self):
        entry = FakeEntry(show="*")
        button = FakeButton()

        QuarkGUITool.toggle_cookie_visibility(None, entry, button)
        self.assertEqual(entry.cget("show"), "")
        self.assertEqual(button.options["text"], "隐藏")

        QuarkGUITool.toggle_cookie_visibility(None, entry, button)
        self.assertEqual(entry.cget("show"), "*")
        self.assertEqual(button.options["text"], "显示")


class TestAdClassification(unittest.TestCase):
    """Shared ad cleanup decision rules."""

    def test_core_resource_with_keyword_is_renamed_not_deleted(self):
        action, new_name = classify_file_action("课程【公众号资料】.mp4", 5 * 1024 * 1024, False, ["公众号"])
        self.assertEqual(action, "rename")
        self.assertEqual(new_name, "课程.mp4")

    def test_ad_image_with_keyword_is_deleted(self):
        action, new_name = classify_file_action("扫码进群.png", 500 * 1024, False, ["扫码"])
        self.assertEqual(action, "delete")
        self.assertIsNone(new_name)

    def test_small_promo_image_is_deleted_without_full_keyword(self):
        action, new_name = classify_file_action("资料群.png", 100 * 1024, False, ["公众号"])
        self.assertEqual(action, "delete")
        self.assertIsNone(new_name)

    def test_directory_with_domain_is_renamed_not_deleted(self):
        action, new_name = classify_file_action("课程-cunlove.cn", 0, True, [])
        self.assertEqual(action, "rename")
        self.assertEqual(new_name, "课程")

    def test_clean_file_is_kept(self):
        action, new_name = classify_file_action("课程讲义.pdf", 1024 * 1024, False, ["公众号"])
        self.assertEqual(action, "keep")
        self.assertIsNone(new_name)

    def test_large_ad_image_deleted_with_new_keywords(self):
        # 4.6MB image should be deleted if the keywords match '加微' or '更多课程'
        action, new_name = classify_file_action("【更多课程加微1505764479】.jpg", int(4.6 * 1024 * 1024), False, ["加微"])
        self.assertEqual(action, "delete")
        self.assertIsNone(new_name)
        
        action, new_name = classify_file_action("【更多课程加微1505764479】.jpg", int(4.6 * 1024 * 1024), False, ["更多课程"])
        self.assertEqual(action, "delete")
        self.assertIsNone(new_name)

    def test_small_docx_ad_is_deleted_under_100kb(self):
        # 34.8KB sensitive docx file should be deleted
        action, new_name = classify_file_action("！网盘资源免费大放送.docx", 35 * 1024, False, ["免费大放送"])
        self.assertEqual(action, "delete")
        self.assertIsNone(new_name)

    def test_large_docx_ad_is_renamed_over_100kb(self):
        # 120KB sensitive docx file should be renamed, not deleted
        action, new_name = classify_file_action("！网盘资源免费大放送.docx", 120 * 1024, False, ["免费大放送"])
        self.assertEqual(action, "rename")
        self.assertEqual(new_name, "！网盘资源.docx")

    def test_small_pdf_ad_is_deleted_under_100kb(self):
        # 49.7KB sensitive PDF file should be deleted
        action, new_name = classify_file_action("更多整理.pdf", int(49.7 * 1024), False, ["更多整理"])
        self.assertEqual(action, "delete")
        self.assertIsNone(new_name)

    def test_numeric_ad_image_deleted_under_100kb(self):
        # 90.8KB image with numeric name should be deleted
        action, new_name = classify_file_action("20250727-40652200.jpg", int(90.8 * 1024), False, [])
        self.assertEqual(action, "delete")
        self.assertIsNone(new_name)

    def test_numeric_ad_image_kept_over_100kb(self):
        # 120KB image with numeric name should be kept (avoid false positives for actual resources)
        action, new_name = classify_file_action("20250727-40652200.jpg", 120 * 1024, False, [])
        self.assertEqual(action, "keep")
        self.assertIsNone(new_name)


class TestInterruptibleSleep(unittest.TestCase):
    """Wait helper used by long-running background tasks."""

    def test_sleep_with_stop_exits_when_stop_flag_turns_true(self):
        calls = []

        def fake_sleep(seconds):
            calls.append(seconds)

        def stop_flag():
            return len(calls) >= 1

        completed = sleep_with_stop(5, stop_flag=stop_flag, interval=1, sleep_func=fake_sleep)
        self.assertFalse(completed)
        self.assertEqual(calls, [1])

    def test_sleep_with_stop_completes_without_stop_flag(self):
        calls = []
        completed = sleep_with_stop(2, stop_flag=lambda: False, interval=1, sleep_func=calls.append)
        self.assertTrue(completed)
        self.assertEqual(calls, [1, 1])


class TestBaiduExistsAndMkdir(unittest.TestCase):
    """Verify that exists checks parent directory and prevents duplicate mkdir API calls."""

    @patch("smart_quark.BaiduBatchTransfer.list_directory")
    def test_exists_returns_true_if_item_found_in_parent(self, mock_list_dir):
        mock_list_dir.return_value = [
            {"server_filename": "18", "isdir": 1, "path": "/微信群日更/2026/06/18"}
        ]
        
        transfer = BaiduBatchTransfer("dummy_cookie")
        self.assertTrue(transfer.exists("/微信群日更/2026/06/18"))
        mock_list_dir.assert_called_once_with("/微信群日更/2026/06")

    @patch("smart_quark.BaiduBatchTransfer.list_directory")
    def test_exists_returns_false_if_not_found(self, mock_list_dir):
        mock_list_dir.return_value = [
            {"server_filename": "other", "isdir": 1, "path": "/微信群日更/2026/06/other"}
        ]
        
        transfer = BaiduBatchTransfer("dummy_cookie")
        self.assertFalse(transfer.exists("/微信群日更/2026/06/18"))
        mock_list_dir.assert_called_once_with("/微信群日更/2026/06")

    @patch("smart_quark.BaiduBatchTransfer.exists")
    @patch("requests.Session.post")
    def test_mkdir_skips_api_if_exists(self, mock_post, mock_exists):
        mock_exists.return_value = True
        
        transfer = BaiduBatchTransfer("dummy_cookie")
        res = transfer.mkdir("/微信群日更/2026/06/18")
        
        self.assertEqual(res, {"errno": 0, "msg": "already exists"})
        mock_post.assert_not_called()


class TestWeChatTextReplacement(unittest.TestCase):
    """Ensure WeChat text template replacement correctly replaces old passwords in both URL query param and separate label."""

    def test_replacement_with_both_query_param_and_label(self):
        import re
        raw_url = "https://pan.baidu.com/s/1WITCCzYgd-wp3mflrVYO9A"
        pwd = "wcv2"
        new_link = "https://pan.baidu.com/s/1WITCCzYgd-wp3mflrVYO9A"
        new_pwd = "n3ac"
        
        current_text = (
            "集链接:\n"
            "https://pan.baidu.com/s/1WITCCzYgd-wp3mflrVYO9A?pwd=wcv2\n"
            "提取码: wcv2"
        )
        
        url_pwd_pattern = re.escape(raw_url) + r'([?&]pwd=)' + re.escape(pwd)
        current_text, url_pwd_count = re.subn(url_pwd_pattern, lambda m: new_link + m.group(1) + new_pwd, current_text)
        
        pattern = re.escape(raw_url) + r'((?:[?&]pwd=(?:' + re.escape(pwd) + r'|' + re.escape(new_pwd) + r'))?\s*(?:[^\w\n]*?(?:提取码|密码|提取密码|码|pwd|密)[:：\s=]*?)\s*)' + re.escape(pwd)
        def repl(m):
            return new_link + m.group(1) + new_pwd
        current_text, count = re.subn(pattern, repl, current_text)
        
        expected = (
            "集链接:\n"
            "https://pan.baidu.com/s/1WITCCzYgd-wp3mflrVYO9A?pwd=n3ac\n"
            "提取码: n3ac"
        )
        self.assertEqual(current_text, expected)
        self.assertEqual(url_pwd_count, 1)
        self.assertEqual(count, 1)

    def test_replacement_with_only_label(self):
        import re
        raw_url = "https://pan.baidu.com/s/1WITCCzYgd-wp3mflrVYO9A"
        pwd = "wcv2"
        new_link = "https://pan.baidu.com/s/1WITCCzYgd-wp3mflrVYO9A"
        new_pwd = "n3ac"
        
        current_text = (
            "集链接:\n"
            "https://pan.baidu.com/s/1WITCCzYgd-wp3mflrVYO9A\n"
            "提取码: wcv2"
        )
        
        url_pwd_pattern = re.escape(raw_url) + r'([?&]pwd=)' + re.escape(pwd)
        current_text, url_pwd_count = re.subn(url_pwd_pattern, lambda m: new_link + m.group(1) + new_pwd, current_text)
        
        pattern = re.escape(raw_url) + r'((?:[?&]pwd=(?:' + re.escape(pwd) + r'|' + re.escape(new_pwd) + r'))?\s*(?:[^\w\n]*?(?:提取码|密码|提取密码|码|pwd|密)[:：\s=]*?)\s*)' + re.escape(pwd)
        def repl(m):
            return new_link + m.group(1) + new_pwd
        current_text, count = re.subn(pattern, repl, current_text)
        
        expected = (
            "集链接:\n"
            "https://pan.baidu.com/s/1WITCCzYgd-wp3mflrVYO9A\n"
            "提取码: n3ac"
        )
        self.assertEqual(current_text, expected)
        self.assertEqual(url_pwd_count, 0)
        self.assertEqual(count, 1)


if __name__ == "__main__":
    unittest.main()
